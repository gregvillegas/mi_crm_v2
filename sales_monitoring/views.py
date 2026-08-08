from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Avg, Sum, F
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.db import models
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, legal
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from .models import (
    SalesActivity, ActivityType, ActivityLog, SupervisorReport,
    ActivityReminder, CallActivity, MeetingActivity, EmailActivity,
    ProposalActivity, TaskActivity, ProofOfConcept
)
from .forms import (
    SalesActivityForm, ActivityFilterForm, QuickActivityForm,
    ActivityUpdateForm, SupervisorReviewForm, BulkActivityUpdateForm,
    ReportGenerationForm, CallActivityForm, MeetingActivityForm,
    EmailActivityForm, ProposalActivityForm, TaskActivityForm,
    ProofOfConceptForm
)
from teams.models import Group, TeamMembership, SupervisorCommitment
from users.models import User
from customers.models import Customer


def _resolve_month_context(month_param):
    today = timezone.now().date()
    if month_param:
        try:
            dt = datetime.strptime(month_param, "%Y-%m").date()
            month_start = dt.replace(day=1)
        except Exception:
            month_start = today.replace(day=1)
    else:
        month_start = today.replace(day=1)

    selected_month_str = month_start.strftime("%Y-%m")

    if month_start.month == 12:
        fiscal_year = month_start.year + 1
    else:
        fiscal_year = month_start.year

    fiscal_start = datetime(fiscal_year - 1, 12, 1).date()
    fiscal_end = datetime(fiscal_year, 11, 30).date()

    fiscal_months = []
    curr = fiscal_start
    for _ in range(12):
        fiscal_months.append({
            'date': curr,
            'label': curr.strftime('%b').upper(),
        })
        if curr.month == 12:
            curr = curr.replace(year=curr.year + 1, month=1)
        else:
            curr = curr.replace(month=curr.month + 1)

    return {
        'today': today,
        'month_start': month_start,
        'selected_month': selected_month_str,
        'fiscal_year': fiscal_year,
        'fiscal_start': fiscal_start,
        'fiscal_end': fiscal_end,
        'fiscal_months': fiscal_months,
        'fiscal_year_label': f"FISCAL YEAR {fiscal_year} ({fiscal_start.strftime('%B %Y').upper()} - {fiscal_end.strftime('%B %Y').upper()})",
    }


def _build_group_fiscal_summary_data(group, month_start):
    month_ctx = _resolve_month_context(month_start.strftime("%Y-%m"))
    salespeople = User.objects.filter(
        team_membership__group=group,
        role='salesperson',
        is_active=True
    )

    fiscal_summary = []
    for sp in salespeople:
        try:
            membership = TeamMembership.objects.get(user=sp)
            quota = membership.quota
        except TeamMembership.DoesNotExist:
            quota = 0

        monthly_data = []
        sp_fiscal_deals = SalesFunnel.objects.filter(
            salesperson=sp,
            deal_outcome='won',
            closed_date__gte=month_ctx['fiscal_start'],
            closed_date__lte=month_ctx['fiscal_end']
        )

        total_fiscal_profit = 0
        for m in month_ctx['fiscal_months']:
            m_start = m['date']
            if m_start.month == 12:
                m_end = m_start.replace(year=m_start.year + 1, month=1) - timedelta(days=1)
            else:
                m_end = m_start.replace(month=m_start.month + 1) - timedelta(days=1)

            m_profit = sp_fiscal_deals.filter(
                closed_date__gte=m_start,
                closed_date__lte=m_end
            ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or 0
            monthly_data.append(m_profit)
            total_fiscal_profit += m_profit

        fiscal_summary.append({
            'salesperson': sp,
            'quota': quota * 12,
            'monthly_data': monthly_data,
            'total_profit': total_fiscal_profit
        })

    return {
        'group': group,
        'selected_month': month_ctx['selected_month'],
        'fiscal_year_label': month_ctx['fiscal_year_label'],
        'fiscal_months': month_ctx['fiscal_months'],
        'fiscal_summary': fiscal_summary,
    }


def _write_group_summary_sheet(ws, summary):
    ws.append([summary['group'].name, summary['group'].team.name, summary['selected_month']])
    ws.append([summary['fiscal_year_label']])
    header = ['Salesperson', 'Annual Quota'] + [m['label'] for m in summary['fiscal_months']] + ['Total Profit']
    ws.append(header)

    title_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    title_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='E5E7EB')
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = title_font
        cell.fill = title_fill
    for cell in ws[2]:
        cell.font = title_font
        cell.fill = title_fill
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in summary['fiscal_summary']:
        ws.append([
            row['salesperson'].get_full_name() or row['salesperson'].username,
            float(row['quota']),
            *[float(v) for v in row['monthly_data']],
            float(row['total_profit']),
        ])

    for column in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        for cell in ws[column]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = u'"P"#,##0.00'

    widths = {
        'A': 28, 'B': 16, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12,
        'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 16
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _build_group_fiscal_summary_pdf(summary):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(legal), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Monthly Achievement Summary", styles['Title']),
        Paragraph(f"{summary['group'].name} - {summary['group'].team.name}", styles['Heading2']),
        Paragraph(summary['fiscal_year_label'], styles['Heading3']),
        Spacer(1, 12),
    ]

    table_data = [['Salesperson', 'Annual Quota'] + [m['label'] for m in summary['fiscal_months']] + ['Total Profit']]
    for item in summary['fiscal_summary']:
        table_data.append([
            item['salesperson'].get_full_name() or item['salesperson'].username,
            f"P{item['quota']:,.2f}",
            *[f"P{profit:,.2f}" if profit > 0 else '-' for profit in item['monthly_data']],
            f"P{item['total_profit']:,.2f}",
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer

@login_required
def dashboard(request):
    """Main dashboard for sales activity monitoring"""
    user = request.user
    
    if user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        return supervisor_dashboard(request)
    elif user.role == 'salesperson':
        return salesperson_dashboard(request)
    elif user.role == 'avp':
        return avp_dashboard(request)
    elif user.role in ['admin', 'vp', 'gm', 'president']:
        return admin_dashboard(request)
    else:
        # Default fallback for other roles
        return admin_dashboard(request)

@login_required
def supervisor_dashboard(request):
    """Dashboard for supervisors, ASMs, and teamleads to monitor their team's activities"""
    user = request.user
    
    if user.role not in ['supervisor', 'asm', 'sm', 'teamlead']:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Get supervised groups based on user role
    if user.role == 'teamlead':
        # Teamleads access groups through led_groups relationship
        supervised_groups = user.led_groups.all()
    elif user.role == 'asm':
        supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
    elif user.role == 'sm':
        supervised_groups = user.sm_groups.all()
    else:
        # Supervisors and ASMs access groups through managed_groups relationship
        supervised_groups = user.managed_groups.all()
    
    # Get all salespeople in supervised groups
    salesperson_ids = []
    for group in supervised_groups:
        salesperson_ids.extend(
            group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
        )
    
    salespeople = User.objects.filter(id__in=salesperson_ids, is_active=True)
    
    # Filter form
    filter_form = ActivityFilterForm(request.GET, supervisor_user=user)
    
    # Base queryset for activities
    activities = SalesActivity.objects.filter(
        salesperson_id__in=salesperson_ids
    ).select_related('salesperson', 'customer', 'activity_type')
    
    # Apply filters
    if filter_form.is_valid():
        cd = filter_form.cleaned_data
        
        if cd['date_from']:
            activities = activities.filter(scheduled_start__date__gte=cd['date_from'])
        if cd['date_to']:
            activities = activities.filter(scheduled_start__date__lte=cd['date_to'])
        if cd['status']:
            activities = activities.filter(status=cd['status'])
        if cd['priority']:
            activities = activities.filter(priority=cd['priority'])
        if cd['activity_type']:
            activities = activities.filter(activity_type=cd['activity_type'])
        if cd['salesperson']:
            activities = activities.filter(salesperson=cd['salesperson'])
        if cd['reviewed_only']:
            activities = activities.filter(reviewed_by_supervisor=True)
        if cd['overdue_only']:
            now = timezone.now()
            activities = activities.filter(
                scheduled_end__lt=now,
                status__in=['planned', 'in_progress']
            )
    
    # Pagination
    paginator = Paginator(activities.order_by('-scheduled_start'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    
    stats = {
        'total_salespeople': salespeople.count(),
        'total_activities': activities.count(),
        'completed_today': activities.filter(
            status='completed',
            actual_end__date=today
        ).count(),
        'overdue_activities': activities.filter(
            scheduled_end__lt=timezone.now(),
            status__in=['planned', 'in_progress']
        ).count(),
        'pending_review': activities.filter(reviewed_by_supervisor=False).count(),
        'this_week_activities': activities.filter(scheduled_start__date__gte=week_start).count(),
        'engineering_required_meetings': activities.filter(
            engineer_required=True,
            activity_type__name__in=['Client Meeting', 'Meeting', 'Client Call']
        ).count(),
    }
    
    # Activity type breakdown
    activity_breakdown = activities.values(
        'activity_type__name', 'activity_type__icon', 'activity_type__color'
    ).annotate(count=Count('id')).order_by('-count')
    
    context = {
        'activities': page_obj,
        'filter_form': filter_form,
        'stats': stats,
        'activity_breakdown': activity_breakdown,
        'supervised_groups': supervised_groups,
        'salespeople': salespeople,
    }
    
    return render(request, 'sales_monitoring/supervisor_dashboard.html', context)

@login_required
def salesperson_dashboard(request):
    """Dashboard for salespeople to view and manage their own activities"""
    user = request.user
    
    if user.role != 'salesperson':
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Get user's activities
    activities = SalesActivity.objects.filter(salesperson=user).select_related(
        'customer', 'activity_type'
    )
    
    # Today's activities
    today = timezone.now().date()
    today_activities = activities.filter(scheduled_start__date=today)
    
    # Upcoming activities (next 7 days)
    next_week = today + timedelta(days=7)
    upcoming_activities = activities.filter(
        scheduled_start__date__gt=today,
        scheduled_start__date__lte=next_week,
        status__in=['planned', 'in_progress']
    ).order_by('scheduled_start')
    
    # Overdue activities
    overdue_activities = activities.filter(
        scheduled_end__lt=timezone.now(),
        status__in=['planned', 'in_progress']
    )
    
    # Recent activity logs
    recent_logs = ActivityLog.objects.filter(
        activity__salesperson=user
    ).select_related('activity').order_by('-timestamp')[:10]
    
    # Statistics
    this_week_start = today - timedelta(days=today.weekday())
    
    # Quota stats
    try:
        membership = TeamMembership.objects.get(user=user)
        quota = membership.quota
    except TeamMembership.DoesNotExist:
        quota = 0
        
    profit_data = SalesFunnel.objects.filter(
        salesperson=user, 
        deal_outcome='won'
    ).aggregate(
        total_profit=Sum(F('retail') - F('cost'))
    )
    total_profit = profit_data['total_profit'] or 0
    quota_achievement = (total_profit / quota * 100) if quota > 0 else 0
    
    stats = {
        'total_activities': activities.count(),
        'today_activities': today_activities.count(),
        'completed_today': activities.filter(
            status='completed',
            actual_end__date=today
        ).count(),
        'upcoming_activities': upcoming_activities.count(),
        'overdue_activities': overdue_activities.count(),
        'this_week_completed': activities.filter(
            status='completed',
            actual_end__date__gte=this_week_start
        ).count(),
        'pending_review': activities.filter(reviewed_by_supervisor=False).count(),
        'quota': quota,
        'total_profit': total_profit,
        'quota_achievement': quota_achievement,
    }
    
    context = {
        'today_activities': today_activities,
        'upcoming_activities': upcoming_activities,
        'overdue_activities': overdue_activities,
        'recent_logs': recent_logs,
        'stats': stats,
    }
    
    return render(request, 'sales_monitoring/salesperson_dashboard.html', context)

@login_required
def salesperson_activity_list(request):
    """List all activities for the current salesperson with edit links"""
    user = request.user
    if user.role != 'salesperson':
        return HttpResponseForbidden("You don't have permission to access this page.")
    activities = SalesActivity.objects.filter(salesperson=user).select_related('activity_type', 'customer').order_by('-scheduled_start', '-created_at')
    status = request.GET.get('status') or ''
    if status:
        activities = activities.filter(status=status)
    paginator = Paginator(activities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'sales_monitoring/salesperson_activity_list.html', {
        'activities': page_obj,
        'status': status
    })
@login_required
def avp_dashboard(request):
    """Dashboard for AVPs to view their team's activity metrics"""
    user = request.user
    
    if user.role != 'avp':
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Get AVP's teams and groups
    from teams.models import Team
    user_teams = Team.objects.filter(avp=user)
    groups = Group.objects.filter(team__in=user_teams)
    
    # Get all salespeople in AVP's teams
    salesperson_ids = []
    for group in groups:
        salesperson_ids.extend(
            group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
        )
    
    salespeople = User.objects.filter(id__in=salesperson_ids, is_active=True)
    
    # Get activities for AVP's teams
    activities = SalesActivity.objects.filter(
        salesperson_id__in=salesperson_ids
    ).select_related('salesperson', 'customer', 'activity_type')
    
    # Statistics and month window (supports ?month=YYYY-MM)
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_param = request.GET.get('month')
    if month_param:
        try:
            dt = datetime.strptime(month_param, "%Y-%m").date()
            month_start = dt.replace(day=1)
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(days=1)
        except Exception:
            month_start = today.replace(day=1)
            month_end = today
    else:
        month_start = today.replace(day=1)
        month_end = today
    
    stats = {
        'total_teams': user_teams.count(),
        'total_groups': groups.count(),
        'total_salespeople': salespeople.count(),
        'total_activities': activities.count(),
        'completed_today': activities.filter(
            status='completed',
            actual_end__date=today
        ).count(),
        'activities_this_week': activities.filter(scheduled_start__date__gte=week_start).count(),
        'activities_this_month': activities.filter(scheduled_start__date__gte=month_start, scheduled_start__date__lte=month_end).count(),
        'overdue_activities': activities.filter(
            scheduled_end__lt=timezone.now(),
            status__in=['planned', 'in_progress']
        ).count(),
        'engineering_required_meetings': activities.filter(
            engineer_required=True,
            activity_type__name__in=['Client Meeting', 'Meeting', 'Client Call']
        ).count(),
    }
    
    # Performance by group
    group_performance = []
    for group in groups:
        group_salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        
        if group_salespeople.exists():
            group_activities = activities.filter(
                salesperson__in=group_salespeople
            )
            
            completed_activities = group_activities.filter(status='completed')
            completion_rate = (completed_activities.count() / group_activities.count() * 100) if group_activities.count() > 0 else 0
            
            # Determine performance color
            perf_color = 'success' if completion_rate >= 80 else 'warning' if completion_rate >= 60 else 'danger'
            group_performance.append({
                'group': group,
                'total_activities': group_activities.count(),
                'completed_activities': completed_activities.count(),
                'completion_rate': completion_rate,
                'salespeople_count': group_salespeople.count(),
                'performance_color': perf_color,
            })
    
    # Activity type breakdown
    activity_breakdown = activities.values(
        'activity_type__name', 'activity_type__icon', 'activity_type__color'
    ).annotate(count=Count('id')).order_by('-count')

    # Group achievements: total quota vs actual profit for the month
    group_achievements = []
    from teams.models import RoleMonthlyQuota
    for group in groups:
        supervisor = group.get_manager()
        group_salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        # Total quota: sum of salesperson quotas + supervisor monthly quota (if any)
        total_quota = TeamMembership.objects.filter(group=group).aggregate(total=Sum('quota'))['total'] or 0
        if supervisor:
            sup_quota = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
            if sup_quota:
                total_quota += sup_quota.amount
        # Actual profit for current month
        profit_data = SalesFunnel.objects.filter(
            salesperson__in=group_salespeople,
            deal_outcome='won',
            closed_date__gte=month_start,
            closed_date__lte=month_end
        ).aggregate(total_profit=Sum(F('retail') - F('cost')))
        actual_profit = profit_data['total_profit'] or 0
        progress_pct = (actual_profit / total_quota * 100) if total_quota and total_quota > 0 else 0
        status_color = 'success' if progress_pct >= 80 else 'warning' if progress_pct >= 60 else 'danger'
        group_achievements.append({
            'group': group,
            'team_name': group.team.name,
            'supervisor_name': supervisor.get_full_name() if supervisor else 'Unassigned',
            'total_quota': total_quota,
            'actual_profit': actual_profit,
            'progress_pct': progress_pct,
            'status_color': status_color,
            'supervisor_id': supervisor.id if supervisor else None,
        })
    
    # Supervisor-only achievements: progress against supervisor's own quota
    supervisor_achievements = []
    for group in groups:
        supervisor = group.get_manager()
        if not supervisor:
            continue
        sup_quota_obj = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
        sup_quota = sup_quota_obj.amount if sup_quota_obj else 0
        # Supervisor achievement should reflect only supervisor-owned deals,
        # not the sum of their group's salesperson deals.
        if getattr(supervisor, 'role', None) == 'salesperson':
            profit_data = SalesFunnel.objects.filter(
                salesperson=supervisor,
                deal_outcome='won',
                closed_date__gte=month_start,
                closed_date__lte=month_end
            ).aggregate(total_profit=Sum(F('retail') - F('cost')))
            actual_profit = profit_data['total_profit'] or 0
        else:
            actual_profit = 0
        sup_progress_pct = (actual_profit / sup_quota * 100) if sup_quota and sup_quota > 0 else 0
        sup_status_color = 'success' if sup_progress_pct >= 80 else 'warning' if sup_progress_pct >= 60 else 'danger'
        supervisor_achievements.append({
            'group': group,
            'team_name': group.team.name,
            'supervisor_name': supervisor.get_full_name(),
            'supervisor_quota': sup_quota,
            'actual_profit': actual_profit,
            'progress_pct': sup_progress_pct,
            'status_color': sup_status_color,
            'supervisor_id': supervisor.id,
        })
    
    # Build team cards with ASM/AVP monthly quotas
    team_cards = []
    for team in user_teams:
        asm_quota = 0
        if team.asm:
            asm_q = RoleMonthlyQuota.objects.filter(user=team.asm, month=month_start).first()
            asm_quota = asm_q.amount if asm_q else 0
        avp_q = RoleMonthlyQuota.objects.filter(user=user, month=month_start).first()
        avp_quota = avp_q.amount if avp_q else 0
        team_cards.append({'team': team, 'asm_quota': asm_quota, 'avp_quota': avp_quota})

    context = {
        'stats': stats,
        'group_performance': group_performance,
        'activity_breakdown': activity_breakdown,
        'user_teams': user_teams,
        'team_cards': team_cards,
        'groups': groups,
        'group_achievements': group_achievements,
        'supervisor_achievements': supervisor_achievements,
    }
    
    return render(request, 'sales_monitoring/avp_dashboard.html', context)


@login_required
def avp_group_activities(request, group_id):
    user = request.user
    if user.role != 'avp':
        return HttpResponseForbidden("You don't have permission to access this page.")

    group = get_object_or_404(Group, id=group_id, team__avp=user)
    salespeople = User.objects.filter(team_membership__group=group, role='salesperson', is_active=True)

    activities = SalesActivity.objects.filter(salesperson__in=salespeople).select_related('salesperson', 'customer', 'activity_type')

    filter_form = ActivityFilterForm(request.GET)
    filter_form.fields['salesperson'].queryset = salespeople

    if filter_form.is_valid():
        cd = filter_form.cleaned_data
        if cd['date_from']:
            activities = activities.filter(scheduled_start__date__gte=cd['date_from'])
        if cd['date_to']:
            activities = activities.filter(scheduled_start__date__lte=cd['date_to'])
        if cd['status']:
            activities = activities.filter(status=cd['status'])
        if cd['priority']:
            activities = activities.filter(priority=cd['priority'])
        if cd['activity_type']:
            activities = activities.filter(activity_type=cd['activity_type'])
        if cd['salesperson']:
            activities = activities.filter(salesperson=cd['salesperson'])
        if cd['reviewed_only']:
            activities = activities.filter(reviewed_by_supervisor=True)
        if cd['overdue_only']:
            activities = activities.filter(
                scheduled_end__lt=timezone.now(),
                status__in=['planned', 'in_progress']
            )

    paginator = Paginator(activities.order_by('-scheduled_start'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    stats = {
        'group_name': group.name,
        'total_activities': activities.count(),
        'completed_today': activities.filter(status='completed', actual_end__date=today).count(),
        'overdue_activities': activities.filter(scheduled_end__lt=timezone.now(), status__in=['planned', 'in_progress']).count(),
        'this_week_activities': activities.filter(scheduled_start__date__gte=week_start).count(),
    }

    context = {
        'group': group,
        'activities': page_obj,
        'filter_form': filter_form,
        'stats': stats,
        'salespeople': salespeople,
    }

    return render(request, 'sales_monitoring/avp_group_activities.html', context)

@login_required
def group_fiscal_summary(request, group_id):
    """Detailed view for group fiscal year summary"""
    user = request.user
    
    allowed_roles = ['supervisor', 'asm', 'sm', 'teamlead', 'avp', 'admin', 'vp', 'gm', 'president']
    if user.role not in allowed_roles:
        return HttpResponseForbidden("You don't have permission to access this page.")
        
    group = get_object_or_404(Group, id=group_id)
    
    # Permission check: ensure user can view this group
    if user.role == 'avp':
        if group.team.avp != user:
             return HttpResponseForbidden("You don't have permission to view this group.")
             
    if user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        # Simplified check - in real app might be stricter
        pass 
    
    month_ctx = _resolve_month_context(request.GET.get('month'))
    summary = _build_group_fiscal_summary_data(group, month_ctx['month_start'])
        
    # Determine back link based on role
    if user.role in ['avp', 'admin', 'vp', 'gm', 'president']:
        back_url_name = 'sales_monitoring:team_performance'
        back_label = 'Back to Team Performance'
    else:
        back_url_name = 'sales_monitoring:group_performance'
        back_label = 'Back to Group Performance'

    context = {
        'group': group,
        'selected_month': summary['selected_month'],
        'fiscal_year_label': summary['fiscal_year_label'],
        'fiscal_months': summary['fiscal_months'],
        'fiscal_summary': summary['fiscal_summary'],
        'back_url_name': back_url_name,
        'back_label': back_label,
    }
    
    return render(request, 'sales_monitoring/group_fiscal_summary.html', context)


@login_required
def export_group_fiscal_summary_excel(request, group_id):
    user = request.user
    allowed_roles = ['supervisor', 'asm', 'sm', 'teamlead', 'avp', 'admin', 'vp', 'gm', 'president']
    if user.role not in allowed_roles:
        return HttpResponseForbidden("You don't have permission to access this page.")

    group = get_object_or_404(Group, id=group_id)
    if user.role == 'avp' and group.team.avp != user:
        return HttpResponseForbidden("You don't have permission to view this group.")

    month_ctx = _resolve_month_context(request.GET.get('month'))
    summary = _build_group_fiscal_summary_data(group, month_ctx['month_start'])

    wb = Workbook()
    ws = wb.active
    ws.title = group.name[:31]
    _write_group_summary_sheet(ws, summary)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"group_fiscal_summary_{group.name}_{summary['selected_month']}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_group_fiscal_summary_pdf(request, group_id):
    user = request.user
    allowed_roles = ['supervisor', 'asm', 'sm', 'teamlead', 'avp', 'admin', 'vp', 'gm', 'president']
    if user.role not in allowed_roles:
        return HttpResponseForbidden("You don't have permission to access this page.")

    group = get_object_or_404(Group, id=group_id)
    if user.role == 'avp' and group.team.avp != user:
        return HttpResponseForbidden("You don't have permission to view this group.")

    month_ctx = _resolve_month_context(request.GET.get('month'))
    summary = _build_group_fiscal_summary_data(group, month_ctx['month_start'])
    pdf_buffer = _build_group_fiscal_summary_pdf(summary)

    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    filename = f"group_fiscal_summary_{group.name}_{summary['selected_month']}.pdf".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_team_fiscal_summary_excel(request):
    user = request.user
    if user.role not in ['avp', 'admin', 'vp', 'gm', 'president']:
        return HttpResponseForbidden("You don't have permission to access this page.")

    if user.role == 'avp':
        from teams.models import Team
        groups = Group.objects.filter(team__in=Team.objects.filter(avp=user))
    else:
        groups = Group.objects.all()

    month_ctx = _resolve_month_context(request.GET.get('month'))
    wb = Workbook()
    first_sheet = True

    for group in groups.order_by('team__name', 'name'):
        summary = _build_group_fiscal_summary_data(group, month_ctx['month_start'])
        ws = wb.active if first_sheet else wb.create_sheet(title=group.name[:31])
        ws.title = group.name[:31]
        _write_group_summary_sheet(ws, summary)
        first_sheet = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"team_fiscal_summary_{month_ctx['selected_month']}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def admin_dashboard(request):
    """Dashboard for admins to view system-wide activity metrics"""
    user = request.user
    
    # Allow access for admin roles and also if no specific role restrictions
    # This ensures the system works even for users without specific roles set
    
    # System-wide statistics
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    total_activities = SalesActivity.objects.all()
    
    stats = {
        'total_activities': total_activities.count(),
        'total_salespeople': User.objects.filter(role='salesperson', is_active=True).count(),
        'activities_today': total_activities.filter(scheduled_start__date=today).count(),
        'completed_today': total_activities.filter(
            status='completed',
            actual_end__date=today
        ).count(),
        'activities_this_week': total_activities.filter(scheduled_start__date__gte=week_start).count(),
        'activities_this_month': total_activities.filter(scheduled_start__date__gte=month_start).count(),
        'overdue_activities': total_activities.filter(
            scheduled_end__lt=timezone.now(),
            status__in=['planned', 'in_progress']
        ).count(),
    }
    
    # Performance by team
    team_performance = []
    for group in Group.objects.all():
        group_salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        
        if group_salespeople.exists():
            group_activities = total_activities.filter(
                salesperson__in=group_salespeople
            )
            
            completed_activities = group_activities.filter(status='completed')
            completion_rate = (completed_activities.count() / group_activities.count() * 100) if group_activities.count() > 0 else 0
            
            team_performance.append({
                'group': group,
                'total_activities': group_activities.count(),
                'completed_activities': completed_activities.count(),
                'completion_rate': completion_rate,
                'salespeople_count': group_salespeople.count(),
            })
    
    # Activity type breakdown
    activity_breakdown = total_activities.values(
        'activity_type__name', 'activity_type__icon', 'activity_type__color'
    ).annotate(count=Count('id')).order_by('-count')
    
    context = {
        'stats': stats,
        'team_performance': team_performance,
        'activity_breakdown': activity_breakdown,
    }
    
    return render(request, 'sales_monitoring/admin_dashboard.html', context)

@login_required
def create_activity(request):
    """Create a new sales activity"""
    user = request.user
    
    # Check if customer ID is provided in query params
    customer_id = request.GET.get('customer')
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, pk=customer_id)
    
    if request.method == 'POST':
        form = SalesActivityForm(request.POST, user=user)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.salesperson = form.cleaned_data.get('resolved_salesperson')
            activity.save()

            if activity.customer and (not activity.customer.is_active or activity.customer.auto_inactive_flag):
                reactivated_fields = []
                if not activity.customer.is_active:
                    activity.customer.is_active = True
                    reactivated_fields.append('is_active')
                if activity.customer.auto_inactive_flag:
                    activity.customer.auto_inactive_flag = False
                    reactivated_fields.append('auto_inactive_flag')
                if reactivated_fields:
                    activity.customer.save(update_fields=reactivated_fields)
            
            # Log the activity creation
            ActivityLog.log_activity_change(
                activity=activity,
                action='created',
                description=f'Activity "{activity.title}" was created',
                changed_by=user
            )
            
            messages.success(request, 'Activity created successfully!')
            return redirect('sales_monitoring:activity_detail', pk=activity.pk)
    else:
        initial_data = {}
        if customer:
            initial_data['customer'] = customer
            
        form = SalesActivityForm(initial=initial_data, user=user)
    
    context = {
        'form': form,
        'title': 'Create New Activity',
        'customer': customer
    }
    
    return render(request, 'sales_monitoring/activity_form.html', context)

@login_required
def create_poc(request):
    """Create a new Proof of Concept"""
    user = request.user
    
    # Check if customer ID is provided in query params
    customer_id = request.GET.get('customer')
    customer = None
    if customer_id:
        customer = get_object_or_404(Customer, pk=customer_id)
    
    if request.method == 'POST':
        form = ProofOfConceptForm(request.POST, user=user)
        if form.is_valid():
            poc = form.save()
            messages.success(request, 'Proof of Concept created successfully!')
            return redirect('customer_detail', pk=poc.customer.pk)
    else:
        initial_data = {}
        if customer:
            initial_data['customer'] = customer
            if hasattr(customer, 'salesperson') and customer.salesperson:
                initial_data['salesperson'] = customer.salesperson
            
        form = ProofOfConceptForm(initial=initial_data, user=user)
    
    context = {
        'form': form,
        'title': 'Create Proof of Concept',
        'customer': customer
    }
    
    return render(request, 'sales_monitoring/poc_form.html', context)

@login_required
def activity_detail(request, pk):
    """View detailed information about a specific activity"""
    activity = get_object_or_404(SalesActivity, pk=pk)
    user = request.user
    
    # Permission check
    if user.role == 'salesperson' and activity.salesperson != user:
        return HttpResponseForbidden("You don't have permission to view this activity.")
    elif user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        # Check if activity belongs to supervised team
        if user.role == 'teamlead':
            supervised_groups = user.led_groups.all()
        elif user.role == 'asm':
            supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
        elif user.role == 'sm':
            supervised_groups = user.sm_groups.all()
        else:
            supervised_groups = user.managed_groups.all()
        
        if not supervised_groups.filter(id=activity.salesperson.team_membership.group_id).exists():
            return HttpResponseForbidden("You don't have permission to view this activity.")
    
    # Get related activity details
    call_details = getattr(activity, 'call_details', None)
    meeting_details = getattr(activity, 'meeting_details', None)
    email_details = getattr(activity, 'email_details', None)
    proposal_details = getattr(activity, 'proposal_details', None)
    task_details = getattr(activity, 'task_details', None)
    
    # Activity logs
    activity_logs = activity.logs.all().select_related('changed_by')
    
    # Supervisor/Teamlead review form
    review_form = None
    can_review = user.role in ['supervisor', 'asm', 'sm', 'teamlead']
    if can_review:
        if request.method == 'POST' and 'review_submit' in request.POST:
            review_form = SupervisorReviewForm(request.POST, include_engineer_required=activity.is_client_meeting)
            if review_form.is_valid():
                cd = review_form.cleaned_data
                if cd['mark_as_reviewed']:
                    previous_notes = activity.supervisor_notes
                    previous_engineer_required = activity.engineer_required
                    activity.mark_reviewed_by_supervisor(
                        user,
                        cd['supervisor_notes'],
                        cd.get('engineer_required'),
                    )
                    
                    ActivityLog.log_activity_change(
                        activity=activity,
                        action='reviewed',
                        description='Activity review updated by supervisor',
                        changed_by=user,
                        old_value={
                            'supervisor_notes': previous_notes,
                            'engineer_required': previous_engineer_required,
                        },
                        new_value={
                            'supervisor_notes': activity.supervisor_notes,
                            'engineer_required': activity.engineer_required,
                        }
                    )
                    
                    messages.success(request, 'Activity review saved successfully!')
                    return redirect('sales_monitoring:activity_detail', pk=pk)
        else:
            review_form = SupervisorReviewForm(
                include_engineer_required=activity.is_client_meeting,
                initial={
                    'supervisor_notes': activity.supervisor_notes,
                    'mark_as_reviewed': True,
                    'engineer_required': activity.engineer_required,
                }
            )
    
    context = {
        'activity': activity,
        'call_details': call_details,
        'meeting_details': meeting_details,
        'email_details': email_details,
        'proposal_details': proposal_details,
        'task_details': task_details,
        'activity_logs': activity_logs,
        'review_form': review_form,
        'can_review': can_review,
    }
    
    return render(request, 'sales_monitoring/activity_detail.html', context)

@login_required
def update_activity(request, pk):
    """Update an existing sales activity"""
    activity = get_object_or_404(SalesActivity, pk=pk)
    user = request.user
    
    # Permission check
    if user.role == 'salesperson' and activity.salesperson != user:
        return HttpResponseForbidden("You don't have permission to edit this activity.")
    
    if request.method == 'POST':
        form = ActivityUpdateForm(request.POST, instance=activity)
        if form.is_valid():
            old_status = activity.status
            activity = form.save()
            
            # Log status changes
            if old_status != activity.status:
                ActivityLog.log_activity_change(
                    activity=activity,
                    action='status_changed',
                    description=f'Status changed from {old_status} to {activity.status}',
                    changed_by=user,
                    old_value={'status': old_status},
                    new_value={'status': activity.status}
                )
            
            # Mark as completed if status is completed
            if activity.status == 'completed' and not activity.actual_end:
                activity.actual_end = timezone.now()
                activity.save()
            
            messages.success(request, 'Activity updated successfully!')
            return redirect('sales_monitoring:activity_detail', pk=pk)
    else:
        form = ActivityUpdateForm(instance=activity)
    
    context = {
        'form': form,
        'activity': activity,
        'title': f'Update {activity.title}',
    }
    
    return render(request, 'sales_monitoring/activity_form.html', context)

@login_required
def quick_log_activity(request):
    """Quick form for logging activities"""
    user = request.user
    
    if user.role != 'salesperson':
        return HttpResponseForbidden("Only salespeople can log activities.")
    
    if request.method == 'POST':
        form = QuickActivityForm(request.POST, user=user)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.salesperson = user
            activity.save()
            
            ActivityLog.log_activity_change(
                activity=activity,
                action='created',
                description=f'Quick activity log: "{activity.title}"',
                changed_by=user
            )
            
            messages.success(request, 'Activity logged successfully!')
            return JsonResponse({'success': True, 'activity_id': activity.pk})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = QuickActivityForm(user=user)
    
    context = {
        'form': form,
    }
    
    return render(request, 'sales_monitoring/quick_activity_form.html', context)

from django.db.models import Sum, F
from sales_funnel.models import SalesFunnel
from teams.models import TeamMembership

@login_required
def team_performance(request):
    """View team performance metrics"""
    user = request.user
    
    if user.role not in ['avp', 'admin', 'vp', 'gm', 'president']:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Determine which groups to show
    if user.role == 'avp':
        # AVPs can only see their own team's groups
        from teams.models import Team
        try:
            user_teams = Team.objects.filter(avp=user)
            groups = Group.objects.filter(team__in=user_teams)
        except:
            groups = Group.objects.none()
    else:
        # Admin and higher roles can see all groups
        groups = Group.objects.all()
    
    # Month selector
    month_param = request.GET.get('month')
    today = timezone.now().date()
    if month_param:
        try:
            dt = datetime.strptime(month_param, "%Y-%m").date()
            month_start = dt.replace(day=1)
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(days=1)
        except Exception:
            month_start = today.replace(day=1)
            month_end = today
    else:
        month_start = today.replace(day=1)
        month_end = today
    selected_month_str = month_start.strftime("%Y-%m")

    # Generate performance data for each group
    performance_data = []
    
    for group in groups:
        salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        
        if salespeople.exists():
            activities = SalesActivity.objects.filter(
                salesperson__in=salespeople
            )
            
            week_start = today - timedelta(days=today.weekday())
            
            # Individual salesperson performance
            salesperson_performance = []
            group_quota = 0
            group_profit = 0
            
            for sp in salespeople:
                # Activity metrics
                sp_activities = activities.filter(salesperson=sp)
                sp_completed = sp_activities.filter(status='completed')
                
                # Quota metrics
                try:
                    membership = TeamMembership.objects.get(user=sp)
                    quota = membership.quota
                except TeamMembership.DoesNotExist:
                    quota = 0
                    
                # Calculate profit (Retail - Cost) for WON deals
                profit_data = SalesFunnel.objects.filter(
                    salesperson=sp, 
                    deal_outcome='won',
                    closed_date__gte=month_start,
                    closed_date__lte=month_end
                ).aggregate(
                    total_profit=Sum(F('retail') - F('cost'))
                )
                total_profit = profit_data['total_profit'] or 0
                
                quota_achievement = (total_profit / quota * 100) if quota > 0 else 0
                
                group_quota += quota
                group_profit += total_profit
                
                sp_data = {
                    'salesperson': sp,
                    'total_activities': sp_activities.count(),
                    'completed_activities': sp_completed.count(),
                    'completion_rate': (sp_completed.count() / sp_activities.count() * 100) if sp_activities.count() > 0 else 0,
                    'this_week_activities': sp_activities.filter(scheduled_start__date__gte=week_start).count(),
                    'overdue_activities': sp_activities.filter(
                        scheduled_end__lt=timezone.now(),
                        status__in=['planned', 'in_progress']
                    ).count(),
                    'quota': quota,
                    'total_profit': total_profit,
                    'quota_achievement': quota_achievement,
                }
                salesperson_performance.append(sp_data)
            
            # Group summary
            group_activities = activities.count()
            group_completed = activities.filter(status='completed').count()
            # Include supervisor monthly quota
            from teams.models import RoleMonthlyQuota
            supervisor = group.get_manager()
            if supervisor:
                sup_quota = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
                if sup_quota:
                    group_quota += sup_quota.amount
            
            # Group Profit is already summed up from salespeople above
            # But let's verify if we need to query again or just use the sum
            # The original code re-queried. Let's stick to the sum loop which is efficient enough
            # unless we need to catch deals from non-active salespeople? 
            # Original code queried SalesFunnel.objects.filter(salesperson__in=salespeople...)
            # which matches the loop.
            
            group_achievement = (group_profit / group_quota * 100) if group_quota > 0 else 0
            
            group_data = {
                'group': group,
                'total_activities': group_activities,
                'completed_activities': group_completed,
                'completion_rate': (group_completed / group_activities * 100) if group_activities > 0 else 0,
                'salespeople_count': salespeople.count(),
                'salesperson_performance': salesperson_performance,
                'group_quota': group_quota,
                'group_profit': group_profit,
                'group_achievement': group_achievement,
            }
            
            performance_data.append(group_data)
    
    context = {
        'performance_data': performance_data,
        'selected_month': selected_month_str,
    }
    
    return render(request, 'sales_monitoring/team_performance.html', context)

@login_required
def group_performance(request):
    """View group performance metrics for supervisors, ASMs, and teamleads"""
    user = request.user
    
    if user.role not in ['supervisor', 'asm', 'sm', 'teamlead']:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    # Get groups based on user role
    if user.role == 'teamlead':
        # Teamleads access groups through led_groups relationship
        supervised_groups = user.led_groups.all()
    elif user.role == 'asm':
        supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
    elif user.role == 'sm':
        # SM sees only their explicitly assigned groups
        supervised_groups = user.sm_groups.all()
    else:
        # Supervisors access groups through managed_groups relationship
        supervised_groups = user.managed_groups.all()
    
    # Month selector (YYYY-MM). Default to current month
    month_param = request.GET.get('month')
    today = timezone.now().date()
    if month_param:
        try:
            dt = datetime.strptime(month_param, "%Y-%m").date()
            month_start = dt.replace(day=1)
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(days=1)
        except Exception:
            month_start = today.replace(day=1)
            month_end = today
    else:
        month_start = today.replace(day=1)
        month_end = today
    selected_month_str = month_start.strftime("%Y-%m")
    
    # Calculate Fiscal Year parameters
    # Fiscal Year 2026 = Dec 2025 to Nov 2026
    if month_start.month == 12:
        fiscal_year = month_start.year + 1
    else:
        fiscal_year = month_start.year
        
    fiscal_start = datetime(fiscal_year - 1, 12, 1).date()
    fiscal_end = datetime(fiscal_year, 11, 30).date()
    
    # Generate list of months for the table header
    fiscal_months = []
    curr = fiscal_start
    # We want exactly 12 months
    for _ in range(12):
        fiscal_months.append({
            'date': curr,
            'label': curr.strftime('%b').upper(), # DEC, JAN, etc.
        })
        # Increment month
        if curr.month == 12:
            curr = curr.replace(year=curr.year + 1, month=1)
        else:
            curr = curr.replace(month=curr.month + 1)

    # Generate performance data for each supervised group
    performance_data = []
    
    for group in supervised_groups:
        salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        activities = SalesActivity.objects.filter(salesperson__in=salespeople)
        
        week_start = today - timedelta(days=today.weekday())
        
        # Individual salesperson performance
        salesperson_performance = []
        group_quota = 0
        group_profit = 0
        
        # Fiscal Year Summary Data
        fiscal_summary = []

        for sp in salespeople:
            # Activity metrics
            sp_activities = activities.filter(salesperson=sp)
            sp_completed = sp_activities.filter(status='completed')
            
            # Quota metrics
            try:
                membership = TeamMembership.objects.get(user=sp)
                quota = membership.quota
            except TeamMembership.DoesNotExist:
                quota = 0
                
            # Profit (Retail - Cost) for WON deals in selected month
            profit_data = SalesFunnel.objects.filter(
                salesperson=sp, 
                deal_outcome='won',
                closed_date__gte=month_start,
                closed_date__lte=month_end
            ).aggregate(total_profit=Sum(F('retail') - F('cost')))
            total_profit = profit_data['total_profit'] or 0
            
            quota_achievement = (total_profit / quota * 100) if quota > 0 else 0
            
            group_quota += quota
            group_profit += total_profit
            
            sp_data = {
                'salesperson': sp,
                'total_activities': sp_activities.count(),
                'completed_activities': sp_completed.count(),
                'completion_rate': (sp_completed.count() / sp_activities.count() * 100) if sp_activities.count() > 0 else 0,
                'this_week_activities': sp_activities.filter(scheduled_start__date__gte=week_start).count(),
                'overdue_activities': sp_activities.filter(
                    scheduled_end__lt=timezone.now(),
                    status__in=['planned', 'in_progress']
                ).count(),
                'quota': quota,
                'total_profit': total_profit,
                'quota_achievement': quota_achievement,
            }
            salesperson_performance.append(sp_data)

            # Calculate monthly data for fiscal year summary
            monthly_data = []
            sp_fiscal_deals = SalesFunnel.objects.filter(
                salesperson=sp,
                deal_outcome='won',
                closed_date__gte=fiscal_start,
                closed_date__lte=fiscal_end
            )
            
            for m in fiscal_months:
                m_start = m['date']
                if m_start.month == 12:
                     m_end = m_start.replace(year=m_start.year + 1, month=1) - timedelta(days=1)
                else:
                     m_end = m_start.replace(month=m_start.month + 1) - timedelta(days=1)
                
                m_profit = sp_fiscal_deals.filter(
                    closed_date__gte=m_start,
                    closed_date__lte=m_end
                ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or 0
                monthly_data.append(m_profit)
            
            fiscal_summary.append({
                'salesperson': sp,
                'quota': quota,
                'monthly_data': monthly_data
            })
        
        # Include supervisor monthly quota
        supervisor = group.get_manager()
        from teams.models import RoleMonthlyQuota
        if supervisor:
            sup_quota = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
            if sup_quota:
                group_quota += sup_quota.amount
        
        # Group summary
        group_activities = activities.count()
        group_completed = activities.filter(status='completed').count()
        group_achievement = (group_profit / group_quota * 100) if group_quota > 0 else 0
        
        group_data = {
            'group': group,
            'total_activities': group_activities,
            'completed_activities': group_completed,
            'completion_rate': (group_completed / group_activities * 100) if group_activities > 0 else 0,
            'salespeople_count': salespeople.count(),
            'salesperson_performance': salesperson_performance,
            'group_quota': group_quota,
            'group_profit': group_profit,
            'group_achievement': group_achievement,
            'fiscal_summary': fiscal_summary,
            'fiscal_year_label': f"FISCAL YEAR {fiscal_year} ({fiscal_start.strftime('%B %Y').upper()} - {fiscal_end.strftime('%B %Y').upper()})",
            'fiscal_months': fiscal_months,
        }
        
        performance_data.append(group_data)
    
    context = {
        'performance_data': performance_data,
        'selected_month': selected_month_str,
    }
    
    return render(request, 'sales_monitoring/group_performance.html', context)

@login_required
def activity_reports(request):
    """Generate and view activity reports"""
    user = request.user
    
    if user.role not in ['avp', 'supervisor', 'asm', 'teamlead', 'admin', 'vp', 'gm', 'president']:
        return HttpResponseForbidden("You don't have permission to access this page.")
    
    if request.method == 'POST':
        form = ReportGenerationForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            
            # Determine report period
            now = timezone.now()
            if cd['report_type'] == 'daily':
                period_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                period_end = period_start + timedelta(days=1)
            elif cd['report_type'] == 'weekly':
                period_start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
                period_end = period_start + timedelta(days=7)
            elif cd['report_type'] == 'monthly':
                period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                next_month = period_start + timedelta(days=32)
                period_end = next_month.replace(day=1)
            else:  # custom
                period_start = cd['period_start']
                period_end = cd['period_end']
            
            # Generate report
            report_data = generate_activity_report(
                user, period_start, period_end, cd['include_individual_breakdown']
            )
            
            context = {
                'report_data': report_data,
                'period_start': period_start,
                'period_end': period_end,
                'report_type': cd['report_type'],
            }
            
            return render(request, 'sales_monitoring/activity_report.html', context)
    else:
        form = ReportGenerationForm()
    
    # Show recent reports
    recent_reports = SupervisorReport.objects.all().order_by('-created_at')
    if user.role == 'supervisor':
        recent_reports = recent_reports.filter(supervisor=user)
    recent_reports = recent_reports[:10]
    
    context = {
        'form': form,
        'recent_reports': recent_reports,
    }
    
    return render(request, 'sales_monitoring/reports.html', context)

def generate_activity_report(user, period_start, period_end, include_breakdown=True):
    """Generate activity report data"""
    
    # Determine scope based on user role
    if user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        # Get groups based on user role
        if user.role == 'teamlead':
            supervised_groups = user.led_groups.all()
        elif user.role == 'asm':
            supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
        elif user.role == 'sm':
            supervised_groups = user.sm_groups.all()
        else:
            supervised_groups = user.managed_groups.all()
        
        salesperson_ids = []
        for group in supervised_groups:
            salesperson_ids.extend(
                group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
            )
        activities_qs = SalesActivity.objects.filter(salesperson_id__in=salesperson_ids)
    elif user.role == 'avp':
        # AVPs can only see activities from their own team's salespeople
        from teams.models import Team
        try:
            user_teams = Team.objects.filter(avp=user)
            team_groups = Group.objects.filter(team__in=user_teams)
            salesperson_ids = []
            for group in team_groups:
                salesperson_ids.extend(
                    group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
                )
            activities_qs = SalesActivity.objects.filter(salesperson_id__in=salesperson_ids)
        except:
            activities_qs = SalesActivity.objects.none()
    else:
        activities_qs = SalesActivity.objects.all()
    
    # Filter by period
    activities = activities_qs.filter(
        created_at__gte=period_start,
        created_at__lte=period_end
    ).select_related('salesperson', 'customer', 'activity_type')
    
    # Summary statistics
    total_activities = activities.count()
    completed_activities = activities.filter(status='completed').count()
    completion_rate = (completed_activities / total_activities * 100) if total_activities > 0 else 0
    
    # Activity type breakdown
    type_breakdown_raw = activities.values(
        'activity_type__name'
    ).annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed'))
    ).order_by('-total')
    
    # Calculate completion rates for each activity type
    type_breakdown = []
    for type_data in type_breakdown_raw:
        completion_rate = (type_data['completed'] / type_data['total'] * 100) if type_data['total'] > 0 else 0
        type_breakdown.append({
            'activity_type__name': type_data['activity_type__name'],
            'total': type_data['total'],
            'completed': type_data['completed'],
            'completion_rate': completion_rate
        })
    
    # Individual breakdown if requested
    individual_breakdown = []
    if include_breakdown:
        salespeople = User.objects.filter(
            sales_activities__in=activities
        ).distinct().order_by('username')
        
        for sp in salespeople:
            sp_activities = activities.filter(salesperson=sp)
            sp_completed = sp_activities.filter(status='completed')
            
            individual_breakdown.append({
                'salesperson': sp,
                'total_activities': sp_activities.count(),
                'completed_activities': sp_completed.count(),
                'completion_rate': (sp_completed.count() / sp_activities.count() * 100) if sp_activities.count() > 0 else 0,
                'activity_breakdown': sp_activities.values(
                    'activity_type__name'
                ).annotate(count=Count('id')).order_by('-count')
            })
    
    return {
        'summary': {
            'total_activities': total_activities,
            'completed_activities': completed_activities,
            'completion_rate': completion_rate,
        },
        'type_breakdown': type_breakdown,
        'individual_breakdown': individual_breakdown,
        'activities': activities,
    }

@login_required
def bulk_update_activities(request):
    """Bulk update multiple activities"""
    user = request.user
    
    if user.role not in ['supervisor', 'asm', 'sm', 'teamlead']:
        return HttpResponseForbidden("You don't have permission to perform bulk updates.")
    
    if request.method == 'POST':
        form = BulkActivityUpdateForm(request.POST, supervisor_user=user)
        if form.is_valid():
            cd = form.cleaned_data
            activities = cd['activities']
            
            updated_count = 0
            for activity in activities:
                if cd['new_status']:
                    old_status = activity.status
                    activity.status = cd['new_status']
                    
                    ActivityLog.log_activity_change(
                        activity=activity,
                        action='status_changed',
                        description=f'Bulk update: Status changed from {old_status} to {activity.status}',
                        changed_by=user
                    )
                
                if cd['mark_as_reviewed']:
                    activity.mark_reviewed_by_supervisor(user, cd['supervisor_notes'])
                    
                    ActivityLog.log_activity_change(
                        activity=activity,
                        action='reviewed',
                        description='Bulk review by supervisor',
                        changed_by=user
                    )
                
                activity.save()
                updated_count += 1
            
            messages.success(request, f'Successfully updated {updated_count} activities!')
            return redirect('sales_monitoring:supervisor_dashboard')
    else:
        form = BulkActivityUpdateForm(supervisor_user=user)
    
    context = {
        'form': form,
        'title': 'Bulk Update Activities',
    }
    
    return render(request, 'sales_monitoring/bulk_update_form.html', context)

@login_required
def activity_calendar(request):
    """Calendar view of activities with week/month switching"""
    user = request.user
    
    # Determine which activities to show based on user role
    if user.role == 'salesperson':
        activities = SalesActivity.objects.filter(salesperson=user)
    elif user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        # Get groups based on user role
        if user.role == 'teamlead':
            supervised_groups = user.led_groups.all()
        elif user.role == 'asm':
            supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
        elif user.role == 'sm':
            supervised_groups = user.sm_groups.all()
        else:
            supervised_groups = user.managed_groups.all()
        
        salesperson_ids = []
        for group in supervised_groups:
            salesperson_ids.extend(
                group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
            )
        activities = SalesActivity.objects.filter(salesperson_id__in=salesperson_ids)
    else:
        activities = SalesActivity.objects.all()
    
    # Determine window from query params
    today = timezone.now().date()
    q_view = (request.GET.get('view') or 'month').lower()
    q_date = request.GET.get('date')
    base_date = today
    try:
        if q_date:
            # Support YYYY-MM-DD or YYYY-MM
            base_date = datetime.strptime(q_date, '%Y-%m-%d').date()
        else:
            base_date = today
    except Exception:
        base_date = today

    if q_view == 'week':
        week_start = base_date - timedelta(days=base_date.weekday())
        week_end = week_start + timedelta(days=7)
        initial_view = 'timeGridWeek'
        initial_date = week_start.isoformat()
        activities = activities.filter(
            scheduled_start__date__gte=week_start,
            scheduled_start__date__lt=week_end
        )
    else:
        month_start = base_date.replace(day=1)
        next_month = (month_start + timedelta(days=32)).replace(day=1)
        initial_view = 'dayGridMonth'
        initial_date = month_start.isoformat()
        activities = activities.filter(
            scheduled_start__date__gte=month_start,
            scheduled_start__date__lt=next_month
        )

    activities = activities.select_related('salesperson', 'customer', 'activity_type')
    
    # Prepare calendar data
    calendar_events = []
    for activity in activities:
        event = {
            'id': activity.id,
            'title': activity.title,
            'start': activity.scheduled_start.isoformat() if activity.scheduled_start else None,
            'end': activity.scheduled_end.isoformat() if activity.scheduled_end else None,
            'color': get_activity_color(activity),
            'url': f'/sales-monitoring/activity/{activity.id}/',
            'extendedProps': {
                'salesperson': activity.salesperson.username,
                'customer': activity.customer.company_name if activity.customer else 'No Customer',
                'status': activity.status,
                'priority': activity.priority,
                'type': activity.activity_type.name,
            }
        }
        calendar_events.append(event)
    
    context = {
        'calendar_events': calendar_events,
        'initial_view': initial_view,
        'initial_date': initial_date,
    }
    
    return render(request, 'sales_monitoring/activity_calendar.html', context)

def get_activity_color(activity):
    """Get color for calendar events based on activity status and priority"""
    if activity.status == 'completed':
        return '#28a745'  # Green
    elif activity.status == 'cancelled':
        return '#6c757d'  # Gray
    elif activity.is_overdue:
        return '#dc3545'  # Red
    elif activity.priority == 'urgent':
        return '#fd7e14'  # Orange
    elif activity.priority == 'high':
        return '#ffc107'  # Yellow
    else:
        return '#007bff'  # Blue

@login_required
def executive_dashboard(request):
    """Executive CRM Dashboard for VP-level view"""
    user = request.user
    
    # Restrict access to VP and higher roles
    if user.role not in ['vp', 'gm', 'president', 'admin']:
        return HttpResponseForbidden("You don't have permission to access the executive dashboard.")
    
    # Get all dashboard data
    dashboard_data = get_executive_dashboard_data()
    
    # Serialize group performance data for JavaScript
    import json
    group_performance_json = json.dumps(dashboard_data['group_performance'])
    
    context = {
        'team_kpis': dashboard_data['team_kpis'],
        'group_performance': dashboard_data['group_performance'],
        'group_performance_json': group_performance_json,
        'funnel_overview': dashboard_data['funnel_overview'],
        'individual_performance': dashboard_data['individual_performance'],
        'insights': dashboard_data['insights'],
        'group_achievements': dashboard_data['group_achievements'],
        'supervisor_achievements': dashboard_data['supervisor_achievements'],
        'company_target': dashboard_data['company_target'],
        'last_updated': timezone.now(),
    }
    
    return render(request, 'sales_monitoring/executive_dashboard.html', context)

def get_executive_dashboard_data():
    """Calculate all data for executive dashboard"""
    from decimal import Decimal
    from sales_funnel.models import SalesFunnel
    from django.db.models import Sum, Avg, Max, Min
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    quarter_start = today.replace(month=((today.month - 1) // 3) * 3 + 1, day=1)
    year_start = today.replace(month=1, day=1)
    
    # 1. Team A vs Team B Performance KPIs (Top-Level)
    from teams.models import Team, TeamMembership
    
    all_funnel_entries = SalesFunnel.objects.all()
    won_deals = all_funnel_entries.filter(deal_outcome='won')
    active_pipeline = all_funnel_entries.filter(is_active=True, deal_outcome='active')
    
    # Get Team A and Team B
    try:
        team_a = Team.objects.get(name='TEAM A')
        team_b = Team.objects.get(name='TEAM B')
    except Team.DoesNotExist:
        # Fallback if teams don't exist
        team_a = Team.objects.first()
        team_b = Team.objects.last()
    
    def get_team_metrics(team):
        # Get all salespeople in this team
        team_salespeople = User.objects.filter(
            team_membership__group__team=team,
            role='salesperson',
            is_active=True
        )
        
        # Calculate current month metrics
        team_current_revenue = won_deals.filter(
            salesperson__in=team_salespeople,
            closed_date__gte=month_start,
            closed_date__lte=today
        ).aggregate(total=Sum('retail'))['total'] or Decimal('0')
        
        team_current_deals = won_deals.filter(
            salesperson__in=team_salespeople,
            closed_date__gte=month_start,
            closed_date__lte=today
        ).count()
        
        # Calculate previous month for growth
        prev_month_start = (month_start - timedelta(days=1)).replace(day=1)
        prev_month_end = month_start - timedelta(days=1)
        
        team_prev_revenue = won_deals.filter(
            salesperson__in=team_salespeople,
            closed_date__gte=prev_month_start,
            closed_date__lte=prev_month_end
        ).aggregate(total=Sum('retail'))['total'] or Decimal('0')
        
        team_prev_deals = won_deals.filter(
            salesperson__in=team_salespeople,
            closed_date__gte=prev_month_start,
            closed_date__lte=prev_month_end
        ).count()
        
        # Calculate pipeline value
        team_pipeline_value = active_pipeline.filter(
            salesperson__in=team_salespeople
        ).aggregate(total=Sum('retail'))['total'] or Decimal('0')
        
        team_active_deals = active_pipeline.filter(
            salesperson__in=team_salespeople
        ).count()
        
        # Calculate growth rates
        revenue_growth = 0
        deals_growth = 0
        if team_prev_revenue > 0:
            revenue_growth = ((team_current_revenue - team_prev_revenue) / team_prev_revenue) * 100
        if team_prev_deals > 0:
            deals_growth = ((team_current_deals - team_prev_deals) / team_prev_deals) * 100
        
        # ASM target commitment (per team)
        asm_target = 0
        if team.asm:
            from teams.models import AsmPersonalTarget
            asm_target_obj = AsmPersonalTarget.objects.filter(team=team, asm=team.asm, month=month_start).first()
            asm_target = asm_target_obj.target_amount if asm_target_obj else 0
        # ASM/AVP monthly quotas
        from teams.models import RoleMonthlyQuota
        asm_quota = 0
        if team.asm:
            asm_q = RoleMonthlyQuota.objects.filter(user=team.asm, month=month_start).first()
            asm_quota = asm_q.amount if asm_q else 0
        avp_q = RoleMonthlyQuota.objects.filter(user=team.avp, month=month_start).first()
        avp_quota = avp_q.amount if avp_q else 0
        # Personal profits for ASM/AVP (if they own deals)
        asm_revenue_this_month = Decimal('0')
        avp_revenue_this_month = Decimal('0')
        if team.asm:
            asm_revenue_this_month = won_deals.filter(
                salesperson=team.asm,
                closed_date__gte=month_start,
                closed_date__lte=today
            ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or Decimal('0')
        if team.avp:
            avp_revenue_this_month = won_deals.filter(
                salesperson=team.avp,
                closed_date__gte=month_start,
                closed_date__lte=today
            ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or Decimal('0')
        # Progress calculation similar to group achievements
        asm_progress_pct = float((asm_revenue_this_month / asm_quota * 100) if asm_quota and asm_quota > 0 else 0)
        avp_progress_pct = float((avp_revenue_this_month / avp_quota * 100) if avp_quota and avp_quota > 0 else 0)
        asm_status_color = 'success' if asm_progress_pct >= 80 else 'warning' if asm_progress_pct >= 60 else 'danger'
        avp_status_color = 'success' if avp_progress_pct >= 80 else 'warning' if avp_progress_pct >= 60 else 'danger'
        return {
            'name': team.name,
            'current_month_revenue': team_current_revenue,
            'revenue_growth': revenue_growth,
            'current_month_deals': team_current_deals,
            'deals_growth': deals_growth,
            'pipeline_value': team_pipeline_value,
            'active_deals': team_active_deals,
            'salesperson_count': team_salespeople.count(),
            'asm_name': team.asm.get_full_name() if team.asm else None,
            'asm_target': asm_target,
            'team_id': team.id,
            'asm_id': team.asm.id if team.asm else None,
            'avp_id': team.avp.id if team.avp else None,
            'asm_quota': asm_quota,
            'avp_quota': avp_quota,
            'asm_progress_pct': asm_progress_pct,
            'avp_progress_pct': avp_progress_pct,
            'asm_status_color': asm_status_color,
            'avp_status_color': avp_status_color,
        }
    
    team_kpis = {
        'team_a': get_team_metrics(team_a),
        'team_b': get_team_metrics(team_b),
    }
    
    # 2. Group Performance (Middle-Level)
    from teams.models import Group, TeamMembership, RoleMonthlyQuota
    
    group_performance = []
    for group in Group.objects.all():
        # Get salespeople in this group
        group_salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        
        if group_salespeople.exists():
            # Get funnel entries for this group
            group_pipeline = all_funnel_entries.filter(
                salesperson__in=group_salespeople,
                is_active=True
            )
            
            group_revenue = all_funnel_entries.filter(
                salesperson__in=group_salespeople,
                deal_outcome='won'
            ).aggregate(total=Sum('retail'))['total'] or Decimal('0')
            
            group_pipeline_value = group_pipeline.aggregate(total=Sum('retail'))['total'] or Decimal('0')
            
            group_performance.append({
                'group_name': group.name,
                'group_type': group.get_group_type_display(),
                'revenue': float(group_revenue),
                'pipeline_value': float(group_pipeline_value),
                'deal_count': group_pipeline.count(),
                'salesperson_count': group_salespeople.count(),
                'team_name': group.team.name,
            })
    
    # Sort by pipeline value descending
    group_performance.sort(key=lambda x: x['pipeline_value'], reverse=True)

    group_achievements = []
    for group in Group.objects.all():
        supervisor = group.get_manager()
        group_salespeople = User.objects.filter(
            team_membership__group=group,
            role='salesperson',
            is_active=True
        )
        actual_profit = won_deals.filter(
            salesperson__in=group_salespeople,
            closed_date__gte=month_start,
            closed_date__lte=today
        ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or Decimal('0')
        # Total quota: sum of salesperson quotas + supervisor monthly quota (if any)
        total_quota = TeamMembership.objects.filter(group=group).aggregate(total=Sum('quota'))['total'] or Decimal('0')
        if supervisor:
            sup_quota = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
            if sup_quota:
                total_quota += sup_quota.amount
        progress_pct = float((actual_profit / total_quota * 100) if total_quota and total_quota > 0 else 0)
        status_color = 'success' if progress_pct >= 80 else 'warning' if progress_pct >= 60 else 'danger'
        group_achievements.append({
            'group_name': group.name,
            'team_name': group.team.name,
            'supervisor_name': supervisor.get_full_name() if supervisor else 'Unassigned',
            'total_quota': float(total_quota),
            'actual_profit': float(actual_profit),
            'progress_pct': progress_pct,
            'status_color': status_color,
        })
    
    # 3. Funnel Overview
    funnel_overview = []
    total_pipeline_value = Decimal('0')
    
    for stage_code, stage_name in SalesFunnel.FUNNEL_STAGES:
        stage_entries = active_pipeline.filter(stage=stage_code)
        stage_value = stage_entries.aggregate(total=Sum('retail'))['total'] or Decimal('0')
        total_pipeline_value += stage_value
        
        # Calculate average age in days
        if stage_entries.exists():
            avg_age = stage_entries.aggregate(
                avg_age=Avg(models.F('updated_at') - models.F('created_at'))
            )['avg_age']
            avg_age_days = avg_age.days if avg_age else 0
        else:
            avg_age_days = 0
        
        # Aging indicator (red if > 30 days, yellow if > 14 days, green otherwise)
        aging_status = 'danger' if avg_age_days > 30 else 'warning' if avg_age_days > 14 else 'success'
        
        funnel_overview.append({
            'stage_code': stage_code,
            'stage_name': stage_name,
            'deal_count': stage_entries.count(),
            'total_value': float(stage_value),
            'avg_age_days': avg_age_days,
            'aging_status': aging_status,
            'percentage': float((stage_value / total_pipeline_value) * 100) if total_pipeline_value > 0 else 0,
        })
    
    # 4. Individual Performance
    individual_performance = []
    all_salespeople = User.objects.filter(role='salesperson', is_active=True)
    
    for salesperson in all_salespeople:
        # Get funnel data
        sp_deals = all_funnel_entries.filter(salesperson=salesperson)
        sp_won_deals = sp_deals.filter(deal_outcome='won')
        sp_active_deals = sp_deals.filter(is_active=True, deal_outcome='active')
        
        # Get activity data
        sp_activities = SalesActivity.objects.filter(salesperson=salesperson)
        sp_activities_this_month = sp_activities.filter(created_at__gte=month_start)
        sp_completed_activities = sp_activities_this_month.filter(status='completed')
        
        # Calculate metrics
        total_revenue = sp_won_deals.aggregate(total=Sum('retail'))['total'] or Decimal('0')
        pipeline_value = sp_active_deals.aggregate(total=Sum('retail'))['total'] or Decimal('0')
        activity_completion_rate = (sp_completed_activities.count() / sp_activities_this_month.count() * 100) if sp_activities_this_month.count() > 0 else 0
        
        # Get team/group info
        try:
            membership = TeamMembership.objects.get(user=salesperson)
            group_name = membership.group.name
            team_name = membership.group.team.name
        except TeamMembership.DoesNotExist:
            group_name = "Unassigned"
            team_name = "Unassigned"
        
        individual_performance.append({
            'salesperson': salesperson,
            'name': f"{salesperson.first_name} {salesperson.last_name}" if salesperson.first_name else salesperson.username,
            'group_name': group_name,
            'team_name': team_name,
            'total_revenue': float(total_revenue),
            'pipeline_value': float(pipeline_value),
            'active_deals': sp_active_deals.count(),
            'won_deals': sp_won_deals.count(),
            'activities_this_month': sp_activities_this_month.count(),
            'activity_completion_rate': activity_completion_rate,
            'avg_deal_size': float(total_revenue / sp_won_deals.count()) if sp_won_deals.count() > 0 else 0,
        })
    
    # Sort by total revenue descending
    individual_performance.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    # 5. Quick Insights Panel
    insights = generate_executive_insights(
        team_kpis, group_performance, funnel_overview, individual_performance
    )
    
    # Company annual target set by GM/VP
    from teams.models import CompanyAnnualTarget
    company_target_obj = CompanyAnnualTarget.objects.filter(year=today.year).first()
    company_quota = company_target_obj.amount if company_target_obj else Decimal('0')
    company_actual_profit = won_deals.filter(
        closed_date__gte=year_start,
        closed_date__lte=today
    ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or Decimal('0')
    company_progress_pct = float((company_actual_profit / company_quota * 100) if company_quota and company_quota > 0 else 0)
    company_status_color = 'success' if company_progress_pct >= 80 else 'warning' if company_progress_pct >= 60 else 'danger'
    company_deficit = float((company_quota - company_actual_profit) if company_quota > company_actual_profit else 0)
    
    # Supervisor-only achievements for executive view
    supervisor_achievements = []
    for group in Group.objects.all():
        supervisor = group.get_manager()
        if not supervisor:
            continue
        sup_quota_obj = RoleMonthlyQuota.objects.filter(user=supervisor, month=month_start).first()
        sup_quota = sup_quota_obj.amount if sup_quota_obj else Decimal('0')
        # Supervisor's personal actual profit (own deals only)
        if getattr(supervisor, 'role', None) == 'salesperson':
            actual_profit = SalesFunnel.objects.filter(
                salesperson=supervisor,
                deal_outcome='won',
                closed_date__gte=month_start,
                closed_date__lte=today
            ).aggregate(total=Sum(F('retail') - F('cost')))['total'] or Decimal('0')
        else:
            actual_profit = Decimal('0')
        sup_progress_pct = float((actual_profit / sup_quota * 100) if sup_quota and sup_quota > 0 else 0)
        sup_status_color = 'success' if sup_progress_pct >= 80 else 'warning' if sup_progress_pct >= 60 else 'danger'
        supervisor_achievements.append({
            'group_name': group.name,
            'team_name': group.team.name,
            'supervisor_name': supervisor.get_full_name(),
            'supervisor_quota': float(sup_quota),
            'actual_profit': float(actual_profit),
            'progress_pct': sup_progress_pct,
            'status_color': sup_status_color,
            'supervisor_id': supervisor.id,
        })
    
    return {
        'team_kpis': team_kpis,
        'group_performance': group_performance,
        'funnel_overview': funnel_overview,
        'individual_performance': individual_performance,
        'insights': insights,
        'group_achievements': group_achievements,
        'supervisor_achievements': supervisor_achievements,
        'company_target': {
            'quota': float(company_quota),
            'actual_profit': float(company_actual_profit),
            'progress_pct': company_progress_pct,
            'deficit': company_deficit,
            'status_color': company_status_color,
        },
    }

def generate_executive_insights(team_kpis, group_performance, funnel_overview, individual_performance):
    """Generate actionable insights for executives"""
    from decimal import Decimal
    insights = []
    
    # Team A vs Team B comparison insights
    team_a = team_kpis['team_a']
    team_b = team_kpis['team_b']
    
    # Convert to float for safe comparisons
    team_a_revenue = float(team_a['current_month_revenue'])
    team_b_revenue = float(team_b['current_month_revenue'])
    team_a_pipeline = float(team_a['pipeline_value'])
    team_b_pipeline = float(team_b['pipeline_value'])
    
    # Revenue comparison
    if team_a_revenue > 0 and team_b_revenue > 0:
        if team_a_revenue > team_b_revenue * 1.5:
            insights.append({
                'type': 'success',
                'title': f"{team_a['name']} Leading Revenue",
                'message': f"{team_a['name']} has {team_a_revenue / team_b_revenue:.1f}x more revenue than {team_b['name']} this month.",
                'action': f"Share {team_a['name']}'s successful strategies with {team_b['name']}."
            })
        elif team_b_revenue > team_a_revenue * 1.5:
            insights.append({
                'type': 'success',
                'title': f"{team_b['name']} Leading Revenue",
                'message': f"{team_b['name']} has {team_b_revenue / team_a_revenue:.1f}x more revenue than {team_a['name']} this month.",
                'action': f"Share {team_b['name']}'s successful strategies with {team_a['name']}."
            })
    
    # Growth rate insights
    if team_a['revenue_growth'] > 20 and team_b['revenue_growth'] < 0:
        insights.append({
            'type': 'warning',
            'title': 'Performance Divergence',
            'message': f"{team_a['name']} growing {team_a['revenue_growth']:.1f}% while {team_b['name']} declining {abs(team_b['revenue_growth']):.1f}%.",
            'action': f"Urgent: Analyze what's working for {team_a['name']} and apply to {team_b['name']}."
        })
    elif team_b['revenue_growth'] > 20 and team_a['revenue_growth'] < 0:
        insights.append({
            'type': 'warning',
            'title': 'Performance Divergence',
            'message': f"{team_b['name']} growing {team_b['revenue_growth']:.1f}% while {team_a['name']} declining {abs(team_a['revenue_growth']):.1f}%.",
            'action': f"Urgent: Analyze what's working for {team_b['name']} and apply to {team_a['name']}."
        })
    
    # Pipeline comparison
    if team_a_pipeline > 0 and team_b_pipeline > 0:
        if team_a_pipeline > team_b_pipeline * 2:
            insights.append({
                'type': 'info',
                'title': 'Pipeline Imbalance',
                'message': f"{team_a['name']} has significantly larger pipeline than {team_b['name']}.",
                'action': f"Consider reallocating leads or providing additional support to {team_b['name']}."
            })
        elif team_b_pipeline > team_a_pipeline * 2:
            insights.append({
                'type': 'info',
                'title': 'Pipeline Imbalance',
                'message': f"{team_b['name']} has significantly larger pipeline than {team_a['name']}.",
                'action': f"Consider reallocating leads or providing additional support to {team_a['name']}."
            })
    
    # Team capacity insights
    if team_a['salesperson_count'] != team_b['salesperson_count'] and team_a['salesperson_count'] > 0 and team_b['salesperson_count'] > 0:
        larger_team = team_a if team_a['salesperson_count'] > team_b['salesperson_count'] else team_b
        smaller_team = team_b if team_a['salesperson_count'] > team_b['salesperson_count'] else team_a
        
        # Calculate revenue per salesperson
        larger_team_per_sp = float(larger_team['current_month_revenue']) / larger_team['salesperson_count']
        smaller_team_per_sp = float(smaller_team['current_month_revenue']) / smaller_team['salesperson_count']
        
        if smaller_team_per_sp > larger_team_per_sp * 1.3:
            insights.append({
                'type': 'success',
                'title': 'Efficiency Excellence',
                'message': f"{smaller_team['name']} generating ₱{smaller_team_per_sp:,.0f} per salesperson vs {larger_team['name']}'s ₱{larger_team_per_sp:,.0f}.",
                'action': f"Study {smaller_team['name']}'s efficiency practices for organization-wide implementation."
            })
    
    # Pipeline coverage insights
    total_combined_pipeline = team_a_pipeline + team_b_pipeline
    total_combined_revenue = team_a_revenue + team_b_revenue
    
    if total_combined_revenue > 0 and total_combined_pipeline < total_combined_revenue * 3:
        insights.append({
            'type': 'warning',
            'title': 'Pipeline Coverage Low',
            'message': 'Combined pipeline is less than 3x monthly revenue target.',
            'action': 'Focus on lead generation and early-stage prospect development across both teams.'
        })
    
    # Funnel aging insights
    for stage in funnel_overview:
        if stage['aging_status'] == 'danger' and stage['deal_count'] > 0:
            insights.append({
                'type': 'danger',
                'title': f"{stage['stage_name']} Stage Aging",
                'message': f"Deals in {stage['stage_name']} averaging {stage['avg_age_days']} days.",
                'action': 'Review and accelerate deals to prevent pipeline stagnation.'
            })
    
    # Individual performance insights
    if len(individual_performance) > 0:
        top_performer = individual_performance[0]
        bottom_performers = [p for p in individual_performance if p['total_revenue'] == 0 and p['pipeline_value'] < 100000]
        
        if len(bottom_performers) > 0:
            insights.append({
                'type': 'info',
                'title': 'Performance Gap Identified',
                'message': f"{len(bottom_performers)} salespeople need attention.",
                'action': f"Consider mentoring from top performer {top_performer['name']}."
            })
    
    # Group performance insights
    if len(group_performance) > 1:
        top_group = group_performance[0]
        bottom_group = group_performance[-1]
        
        if top_group['pipeline_value'] > bottom_group['pipeline_value'] * 3:
            insights.append({
                'type': 'info',
                'title': 'Group Performance Disparity',
                'message': f"{top_group['group_name']} significantly outperforming {bottom_group['group_name']}.",
                'action': 'Analyze and replicate successful strategies across groups.'
            })
    
    return insights

@login_required
def export_activities(request):
    """Export activities to CSV"""
    import csv
    from django.http import HttpResponse
    
    user = request.user
    
    # Determine which activities to export based on user role
    if user.role == 'salesperson':
        activities = SalesActivity.objects.filter(salesperson=user)
    elif user.role in ['supervisor', 'asm', 'sm', 'teamlead']:
        # Get groups based on user role
        if user.role == 'teamlead':
            supervised_groups = user.led_groups.all()
        elif user.role == 'asm':
            supervised_groups = Group.objects.filter(team__in=user.asm_teams.all())
        elif user.role == 'sm':
            supervised_groups = user.sm_groups.all()
        else:
            supervised_groups = user.managed_groups.all()
        
        salesperson_ids = []
        for group in supervised_groups:
            salesperson_ids.extend(
                group.members.filter(user__role='salesperson').values_list('user_id', flat=True)
            )
        activities = SalesActivity.objects.filter(salesperson_id__in=salesperson_ids)
    else:
        activities = SalesActivity.objects.all()
    
    activities = activities.select_related('salesperson', 'customer', 'activity_type')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_activities_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Title', 'Activity Type', 'Salesperson', 'Customer', 'Status', 'Priority',
        'Scheduled Start', 'Scheduled End', 'Actual Start', 'Actual End',
        'Duration (minutes)', 'Reviewed by Supervisor', 'Created At'
    ])
    
    for activity in activities:
        writer.writerow([
            activity.id,
            activity.title,
            activity.activity_type.name,
            activity.salesperson.username,
            activity.customer.company_name if activity.customer else '',
            activity.get_status_display(),
            activity.get_priority_display(),
            activity.scheduled_start.strftime('%Y-%m-%d %H:%M') if activity.scheduled_start else '',
            activity.scheduled_end.strftime('%Y-%m-%d %H:%M') if activity.scheduled_end else '',
            activity.actual_start.strftime('%Y-%m-%d %H:%M') if activity.actual_start else '',
            activity.actual_end.strftime('%Y-%m-%d %H:%M') if activity.actual_end else '',
            activity.duration_minutes or '',
            'Yes' if activity.reviewed_by_supervisor else 'No',
            activity.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    
    return response
